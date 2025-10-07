import pandas as pd
from openpyxl import load_workbook

# ----------------------------
# Configurations
# ----------------------------
REQUIRED_ATTRIBUTES = {
    "posted_dt": {"mandatory": False},
    "doc_dt": {"mandatory": False},
    "doc": {"mandatory": False},
    "memo_description": {"mandatory": False},
    "department_name": {"mandatory": False},
    "supplier_name": {"mandatory": False},
    "item_name": {"mandatory": False},
    "customer_name": {"mandatory": False},
    "supplier": {"mandatory": False},
    "jnl": {"mandatory": False},
}

# Sample variations for header name mapping
SAMPLE_GL_HEADERS = [
    "posted_dt", "posting date",
    "doc_dt", "document date",
    "doc", "document number",
    "memo_description", "description",
    "department_name", "department",
    "supplier_name", "vendor",
    "account_name", "account",
    "debit", "credit", "currency", "amount"
]

# ----------------------------
# Helper functions
# ----------------------------
def normalize_header(header):
    """Normalize header names (lowercase, remove extra spaces and special chars)."""
    return str(header).strip().lower().replace(" ", "_").replace("-", "_")

def find_header_row(file_path):
    """Finds the header row dynamically in an Excel file."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row and any(cell for cell in row):
                normalized = [normalize_header(cell) for cell in row if cell]
                # Check if at least 3 known GL headers are found
                matches = sum(
                    1 for h in normalized if any(h in s for s in SAMPLE_GL_HEADERS)
                )
                if matches >= 3:
                    return i, sheet_name
    raise ValueError("Header row not found in any sheet.")

def validate_gl(file_path):
    """Validates the presence of GL attributes in the file."""
    header_row, sheet_name = find_header_row(file_path)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    normalized_cols = [normalize_header(col) for col in df.columns]

    missing = []
    for attr in REQUIRED_ATTRIBUTES.keys():
        if not any(attr in col for col in normalized_cols):
            missing.append(attr)

    print(f"Sheet: {sheet_name}")
    print(f"Detected Header Row: {header_row + 1}")
    print("Available Columns:", df.columns.tolist())
    print("\nMissing Attributes:" if missing else "\nAll required attributes found.")
    for m in missing:
        print(f"- {m}")

# ----------------------------
# Example usage
# ----------------------------
if __name__ == "__main__":
    file_path = "path_to_your_gl_file.xlsx"
    validate_gl(file_path)
